import csv, sys, os, re
# Excel xlsx processing
import openpyxl

#########
# Sample program to test out the various functions 
# of the XLSX (Excel) python library. 
#########

# Print out the version of openpy
op_ver = openpyxl.__version__
print("Version=%s" % op_ver)

# Hard code here to quicken the testing
excel_file = "testdir1/she1126-test.xlsx"  # One of the Shepher files
output_file = "testout1/excel-1.csv"

# Open the file up
wb = openpyxl.load_workbook(excel_file)

# Grab all sheet names
sheets = wb.sheetnames
for sht in sheets:
    print("  %s" % sht)
print("First sheet is [%s]" % sheets[0])

# Print # rows/cols
# First get active sheet
active_sheet = wb.active
print("Active ws title: %s:" % active_sheet.title)
print("sheet.max_row = %d" % active_sheet.max_row)
print("sheet.max_column = %d" % active_sheet.max_column)

# Print values of top columns
col_list = []
for col in range(active_sheet.max_column):
    # Must start at 1 w excel
    col_list.append(active_sheet.cell(row=1,column=col+1).value.strip())
print(col_list)

# returns 4 which 5th (0 based)
base_price_col = col_list.index("Base Price")
print("Base price is in column %d" % base_price_col)

# Print first 5 base price values
for arow in active_sheet.iter_rows(max_row=5,min_row=2):
    print("   Available=%d, Price=%s" % (int(arow[0].value), arow[base_price_col].value))

