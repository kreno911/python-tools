import csv, sys, os, re
# Excel xlsx processing
import openpyxl
from mypylib import contains_char

# create_bcr_file.py --input <file>.xlsx > 9_13.csv
# create_bcr_file.py --input <file>.xlsx --sheets (to get sheet names)
# create_bcr_file.py --input <file>.xlsx --upc # --price # (UPC/price columns)
#                                    defaults: G, E

######################################
# This will create the initial csv file to feed into the analyzer
# This will print out price,UPC with header: Base Price,UPC Code 
#    - redirect these to a .csv file
#    - upload to Beaus tool
######################################
def processExcelForBow(sheet):
    print("Base Price,UPC Code")
    for row in range(2, sheet.max_row + 1):
        # Make sure quantity is > 0
        quantity = sheet[quantity_col + str(row)].value
        if check_quantity:
            if '+' in str(quantity):  # Schepher thing '1800+'
                break
            if quantity == "None" or int(quantity) <= 0:
                continue
        # Base price could be different
        #price = sheet['D'+str (row)].value
        price = sheet[price_column + str(row)].value
        if price is None:
            continue
        # Error correction like above
        upc = str(sheet[upc_column + str(row)].value)
        # ACDitro has UPCs with characters
        containsChar = contains_char(upc)
        if upc == "None" or upc == "- None -" or "'" in upc or containsChar:
            #print("skip: " + upc)
            continue
        # Some upc's come in format: 26608001249.0, strip any decimals
        if "." in upc:
            #print("Found: ", upc)
            upc = upc.split(".")[0]
        upc_fixed = "%012d" % int(upc)
        print("%s,%s" % (price,upc_fixed))

bow_file = "none"
sheet_name = "Inventory"
show_sheets = False
# Default columns 
quantity_col = 'A'   # Shepher 
check_quantity = False
upc_column = 'G'
price_column = 'E'
length = len(sys.argv)
if length < 2:
    print("Options: --input <xlsx> [--sheetname <sheet>|--sheets]")
    print("Option --sheets will print sheet names and exit.")
    print("       --upccol <char> -> UPC column (default is 'G')")
    print("       --pricecol <char> Price column (default is 'E')")
    print("       --quantity-col <char> (default is 'A')")
    print("Sheet name default is %s" % sheet_name)
    sys.exit(11)
else:
    for v in range(1, length):
        # Export for Bow's tool (price,upc)
        if sys.argv[v] == "--input":
            bow_file = sys.argv[v+1]
        if sys.argv[v] == "--sheets":
            show_sheets = True
        if sys.argv[v] == "--sheetname":
            sheet_name = sys.argv[v+1]
        if sys.argv[v] == "--upccol":
            upc_column = sys.argv[v+1]
        if sys.argv[v] == "--pricecol":
            price_column = sys.argv[v+1]
        # Some spreadsheets come with many 0 quantity columns, filter them 
        if sys.argv[v] == "--quantity-col":
            quantity_col = sys.argv[v+1]
            check_quantity = True

if bow_file != "none":
    if not os.path.exists(bow_file):
        print("File %s does not exist." % bow_file)
        sys.exit(11)
    wb = openpyxl.load_workbook(bow_file)
    # If we want to show sheets, just print them and exit
    if show_sheets:
        sheets = wb.sheetnames
        for sht in sheets:
            print("-->%s" % sht)
    else:
        # Can access a sheet by name with wb[sheets[<sheet_num>]]
        bow_sheet = wb.active #wb.get_sheet_by_name(sheet_name)
        processExcelForBow(bow_sheet)
