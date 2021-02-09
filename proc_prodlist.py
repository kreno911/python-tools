
import csv, sys, os, re
# Excel xlsx processing 
import openpyxl
from mypylib import contains_char

# Run as: proc_prodlist.py --results 8_27-special_results.xlsx
#           --orig special-8-27-19.xlsx > final_results.csv
# Terminal: View->Tool Windows->Terminal
# Running on windows:
#  \python\Python27\python.exe proc_prodlist.py --orig shepher-9_13.xlsx
#                                               --results shepher-9_13-result.xlsx
#                                               --shtname Inventory

# If you see: 
#   ValueError: invalid literal for int() with base 10: '715471200010.0'
#   Go to the original spreadsheet (xlsx) and format UPC column to number, 0 decimals

######################################
# Return a map of UPC to supplier item #
# Params:
#   sheet - the sheet to evaluate
#   supplier_item_col - column containing the suppliers item #
# Return example (map):
#   { '722626008276': 'BCDDGM80GRY', 
#     '722626007477': 'BCDDGM80ORG' 
#   }
######################################
def getUPCtoSKUMap(sheet, supplier_item_col):
    data = {}
    # Start at row 2 to skip header
    for row in range(2, sheet.max_row + 1):
        upc = str(sheet[upc_column + str(row)].value)
        # ACDitro has UPCs with characters
        containsChar = contains_char(upc)
        if upc == "None" or upc == "- None -" or "'" in upc or containsChar or upc == "":
            #print("skip: " + upc)
            continue
        # UPC fields that start with 0 are truncated
        # so 033 will be 33 which is wrong

        # Did some testing, some UPCs are 13 or 14 chars long but
        # any I added '0' to front matched a leading zero check
        # First exported whole UPC row to CSV
        #if str(upc_fixed)[0] == "0":
            #print(upc_fixed
        upc_fixed = "%012d" % int(upc)
        sku = sheet[supplier_item_col+str(row)].value
        data[upc_fixed] = sku
    return data

######################################
# This will create a master list containing ASIN/UPC/Supplier SKU
# from the original and results spreadsheet. 
# Results file will usually be smaller (less rows) than original
# Original contains:
#   UPC/Supplier Sku/Supplier Desc
# Results contains:
#   ASIN/Profit/Amz title
######################################
def processCombined(result_sheet, orig_sheet):
    print("ASIN,UPC,SupplierSku,Price,AmzDescription")
    # UPC,SKU
    data = getUPCtoSKUMap(orig_sheet, supplier_item_col)
    for row in range(2, result_sheet.max_row + 1):
        price = result_sheet['E'+str(row)].value
        if price is None:
            continue
        profit = result_sheet['I'+str(row)].value
        price = result_sheet['E' + str(row)].value
        asin = result_sheet['A'+str(row)].value
        upc  = result_sheet['B'+str(row)].value.lstrip()
        amz_desc   = result_sheet['D' + str(row)].value
        suppl_code = "none"
        if upc in data:
            suppl_code = data[upc]
        if suppl_code == "none":
            print("Supplier code is NONE!!!! [%s] (Check data assignment)" % upc)
            sys.exit(13)
        # Need to remove spacial characters in description
        amz_desc = re.sub('[^A-Za-z0-9]+', ' ', amz_desc)
        print("%s,%s,%s,%s,%s" % (asin,upc,suppl_code,price,amz_desc))

original_file = "none"
sheet_name = "Inventory"
upc_column = 'G'
supplier_item_col = 'B'
length = len(sys.argv)
if length < 2:
    print("\nUse this to generate a cross list of product UPCs to ASINs to supplier codes")
    print("--shtname <sheet nm>|--orig <xlsx>|--results <xlsx>")
    print("provide both orig/results together")
    print("--shtname is for sheets that are different than default. (Inventory)")
    print("--orig is the original xlsx file")
    sys.exit(11)
else:
    for v in range(1, length):
        if sys.argv[v] == "--shtname":
            sheet_name = sys.argv[v+1]
        # original file we grab UPC/Supplier SKUs
        if sys.argv[v] == "--orig":
            original_file = sys.argv[v+1]
        if sys.argv[v] == "--results":
            results_file = sys.argv[v+1]
        if sys.argv[v] == "--upccol":
            upc_column = sys.argv[v+1]
        if sys.argv[v] == "--supitemcol":
            supplier_item_col = sys.argv[v+1]

if original_file != "none" and results_file != "none":
    wb = openpyxl.load_workbook(original_file)
    orig_sheet = wb.active  
    wb2 = openpyxl.load_workbook(results_file)
    result_sheet = wb2.active 
    processCombined(result_sheet, orig_sheet)

