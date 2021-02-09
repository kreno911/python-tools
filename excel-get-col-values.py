import csv, sys, os, re
# Excel xlsx processing
import openpyxl

######################################
# TEST SCRIPT FOR openpyxl
#
# This will print out column values for a given excel spreadsheet
# This script requires the following parameters:
#   --sheetname -> the name of the sheet in the spreadsheet to use
#   --file -> the Excel spreadsheet 
# Optional:
#   --sheets -> displays the sheet names and exits
######################################

#######
# Just demonstrating how to extract data...
#######
def printColumns(sheet, columns):
    col_data = {}
    limit = 5
    for c in columns:
        print("Column %s data:" % c)
        value_list = []
        count = 0
        row_value_arrays = []
        for row in sheet['I':'K']:
            row_values = (c.value for c in row)
            row_value_arrays.append(row_values)
            # Print ALL row values for this column 
            #print(list(row_values))
        # Just print length
        print(len(row_value_arrays))
        for row in range(2, sheet.max_row + 1):
            value_list.append(sheet[c + str(row)].value)
            # Change limit to like 500 to get all
            if count == limit:
                break
            print(" ", sheet[c + str(row)].value)
            count += 1
        col_data[c] = value_list
    # Print them together    
    # for col in col_data:
    #     alist =  col_data[col]
    #     for value in alist:
    #         print(" :", value)

excel_file = "none"
sheet_name = "none"
show_sheets = False
columns = []
length = len(sys.argv)
for v in range(1, length):
    if sys.argv[v] == "--file":
        excel_file = sys.argv[v+1]
    if sys.argv[v] == "--sheets":
        show_sheets = True
    if sys.argv[v] == "--sheetname":
        sheet_name = sys.argv[v+1]
    if sys.argv[v] == "--columns":
        cols = sys.argv[v+1].split(",")
        columns = map(str.upper, cols)

if excel_file == "none" or not os.path.isfile(excel_file):
    print("Need valid excel file!")
    sys.exit(11)

wb = openpyxl.load_workbook(excel_file)
# If we want to show sheets, just print them and exit
if show_sheets:
    sheets = wb.sheetnames
    for sht in sheets:
        print("-->%s" % sht)
else:
    if sheet_name == "none":
        print("Need a sheet name, use --sheets to display sheetnames")
        sys.exit(12)
    if not columns:
        print("Need valid columns: [a,b,...]")
        sys.exit(12)
    # How to get by name 
    sheet = wb.active #wb.get_sheet_by_name(sheet_name)
    printColumns(sheet, list(columns))
